"""Генерация Excel-отчётов для калькуляторов (openpyxl)."""

import io
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers

from bot.config.rates import (
    EPB,
    INSURANCE_ABOVE,
    INSURANCE_BASE,
    NDFL_SCALE,
    NDFL_SCALE_NORTH,
    TERRITORY_GROUPS,
)

TWO_PLACES = Decimal("0.01")
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TOTAL_FONT = Font(bold=True, size=11, color="1F4E79")
_TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_NUM_FMT = '#,##0.00'


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_total(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        cell.border = _THIN_BORDER


def _style_data(ws, row, max_col, num_cols=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = _THIN_BORDER
        if num_cols and col in num_cols:
            cell.number_format = _NUM_FMT
            cell.alignment = Alignment(horizontal="right")


def _apply_scale(income: int, scale: list) -> Decimal:
    total_tax = Decimal(0)
    prev_bound = 0
    for bound, rate in scale:
        if bound is None:
            taxable = Decimal(income - prev_bound)
            total_tax += taxable * rate
            break
        if income <= bound:
            taxable = Decimal(income - prev_bound)
            total_tax += taxable * rate
            break
        taxable = Decimal(bound - prev_bound)
        total_tax += taxable * rate
        prev_bound = bound
    return total_tax.quantize(TWO_PLACES, ROUND_HALF_UP)


# ─────────────────────────────────────────────
# ЗАРПЛАТА
# ─────────────────────────────────────────────

def export_salary_report(territory: str, oklad: int, nadbavka_pct: int) -> io.BytesIO:
    """Excel-отчёт по расчёту зарплаты."""
    group = TERRITORY_GROUPS.get(territory, {})
    oklad_d = Decimal(oklad)
    rk = group.get("rk", Decimal("1.3"))
    rk_extra = rk - 1
    max_nadb = group.get("max_nadbavka", Decimal(0))
    nadbavka = min(Decimal(nadbavka_pct) / 100, max_nadb)

    rk_sum = (oklad_d * rk_extra).quantize(TWO_PLACES, ROUND_HALF_UP)
    nadb_sum = (oklad_d * nadbavka).quantize(TWO_PLACES, ROUND_HALF_UP)
    gross = oklad_d + rk_sum + nadb_sum

    annual_base = int(oklad_d * 12)
    annual_north = int((rk_sum + nadb_sum) * 12)
    ndfl_base = _apply_scale(annual_base, NDFL_SCALE)
    ndfl_north = _apply_scale(annual_north, NDFL_SCALE_NORTH)
    ndfl_monthly = ((ndfl_base + ndfl_north) / 12).quantize(TWO_PLACES, ROUND_HALF_UP)
    net = gross - ndfl_monthly

    # Страховые взносы работодателя
    insurance = (gross * INSURANCE_BASE).quantize(TWO_PLACES, ROUND_HALF_UP)

    wb = Workbook()
    ws = wb.active
    ws.title = "Расчёт зарплаты"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    headers = ["Показатель", "Сумма, ₽"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, 2)

    rows_data = [
        ("Территория", group.get("name", territory)),
        ("Районный коэффициент", float(rk)),
        ("Северная надбавка, %", int(nadbavka * 100)),
        ("", ""),
        ("Оклад", float(oklad_d)),
        (f"РК ({rk_extra})", float(rk_sum)),
        (f"Надбавка ({int(nadbavka * 100)}%)", float(nadb_sum)),
        ("Начислено (gross)", float(gross)),
        ("", ""),
        ("НДФЛ (основная часть, мес.)", float((ndfl_base / 12).quantize(TWO_PLACES, ROUND_HALF_UP))),
        ("НДФЛ (северная часть, мес.)", float((ndfl_north / 12).quantize(TWO_PLACES, ROUND_HALF_UP))),
        ("НДФЛ итого, мес.", float(ndfl_monthly)),
        ("На руки (net)", float(net)),
        ("", ""),
        ("Страховые взносы (30%), мес.", float(insurance)),
        ("Полная стоимость сотрудника", float(gross + insurance)),
    ]

    for i, (label, value) in enumerate(rows_data, 2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        num_cols = {2} if isinstance(value, float) else None
        _style_data(ws, i, 2, num_cols)

    # Выделяем итоговые строки
    for total_row in [9, 14]:  # Начислено, На руки
        _style_total(ws, total_row, 2)
    _style_total(ws, len(rows_data) + 1, 2)  # Полная стоимость

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# СТРАХОВЫЕ ВЗНОСЫ — ПОМЕСЯЧНАЯ РАЗБИВКА
# ─────────────────────────────────────────────

def export_contributions_report(monthly_salary: int) -> io.BytesIO:
    """Excel с помесячной разбивкой взносов за год."""
    monthly = Decimal(monthly_salary)

    month_names = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Страховые взносы"

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    headers = ["Месяц", "Зарплата, ₽", "Нарастающий итог, ₽",
               "Ставка, %", "Взносы, ₽", "ЕПБ исчерп."]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, 6)

    total_contributions = Decimal(0)
    cumulative = Decimal(0)

    for i, name in enumerate(month_names, 1):
        prev_cumulative = cumulative
        cumulative += monthly
        row = i + 1

        # Определяем ставку и базу
        if prev_cumulative >= EPB:
            # Весь месяц свыше ЕПБ
            contrib = (monthly * INSURANCE_ABOVE).quantize(TWO_PLACES, ROUND_HALF_UP)
            rate = float(INSURANCE_ABOVE * 100)
            exhausted = ""
        elif cumulative > EPB:
            # Месяц пересечения ЕПБ
            within = EPB - prev_cumulative
            above = cumulative - EPB
            contrib = (
                (within * INSURANCE_BASE + above * INSURANCE_ABOVE)
                .quantize(TWO_PLACES, ROUND_HALF_UP)
            )
            rate_text = "переход"
            exhausted = "✓"
            ws.cell(row=row, column=4, value=rate_text)
        else:
            contrib = (monthly * INSURANCE_BASE).quantize(TWO_PLACES, ROUND_HALF_UP)
            rate = float(INSURANCE_BASE * 100)
            exhausted = ""

        total_contributions += contrib

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=float(monthly))
        ws.cell(row=row, column=3, value=float(cumulative))
        if not (prev_cumulative < EPB < cumulative):
            ws.cell(row=row, column=4, value=rate)
        ws.cell(row=row, column=5, value=float(contrib))
        ws.cell(row=row, column=6, value=exhausted)
        _style_data(ws, row, 6, {2, 3, 5})

    # Итого
    total_row = 14
    ws.cell(row=total_row, column=1, value="ИТОГО")
    ws.cell(row=total_row, column=2, value=float(monthly * 12))
    ws.cell(row=total_row, column=5, value=float(total_contributions))
    _style_total(ws, total_row, 6)
    for col in [2, 5]:
        ws.cell(row=total_row, column=col).number_format = _NUM_FMT

    # Доп. инфо
    ws.cell(row=16, column=1, value=f"ЕПБ 2026: {float(EPB):,.0f} ₽")
    ws.cell(row=17, column=1, value=f"Ставка до ЕПБ: {float(INSURANCE_BASE * 100)}%")
    ws.cell(row=18, column=1, value=f"Ставка свыше ЕПБ: {float(INSURANCE_ABOVE * 100)}%")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
# НДФЛ — ДЕТАЛИЗАЦИЯ ПО СТУПЕНЯМ
# ─────────────────────────────────────────────

def export_ndfl_report(annual_income: int) -> io.BytesIO:
    """Excel с детализацией НДФЛ по ступеням прогрессивной шкалы."""
    wb = Workbook()
    ws = wb.active
    ws.title = "НДФЛ 2026"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18

    headers = ["От, ₽", "До, ₽", "Ставка, %", "Облагаемая база, ₽", "НДФЛ, ₽"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, 5)

    total_tax = Decimal(0)
    row = 2
    prev = 0

    for bound, rate in NDFL_SCALE:
        if annual_income <= prev:
            break
        top = bound if bound and annual_income > bound else annual_income
        taxable = Decimal(top - prev)
        part_tax = (taxable * rate).quantize(TWO_PLACES, ROUND_HALF_UP)
        total_tax += part_tax

        ws.cell(row=row, column=1, value=float(prev))
        ws.cell(row=row, column=2, value=float(top))
        ws.cell(row=row, column=3, value=float(rate * 100))
        ws.cell(row=row, column=4, value=float(taxable))
        ws.cell(row=row, column=5, value=float(part_tax))
        _style_data(ws, row, 5, {1, 2, 4, 5})

        if bound is None or annual_income <= bound:
            break
        prev = bound
        row += 1

    # Итого
    row += 1
    ws.cell(row=row, column=1, value="ИТОГО")
    ws.cell(row=row, column=4, value=float(Decimal(annual_income)))
    ws.cell(row=row, column=5, value=float(total_tax))
    _style_total(ws, row, 5)
    for col in [4, 5]:
        ws.cell(row=row, column=col).number_format = _NUM_FMT

    # Сводка
    income_d = Decimal(annual_income)
    effective = (total_tax / income_d * 100).quantize(TWO_PLACES, ROUND_HALF_UP) if income_d else Decimal(0)
    net = income_d - total_tax

    row += 2
    ws.cell(row=row, column=1, value="Годовой доход:")
    ws.cell(row=row, column=2, value=float(income_d))
    ws.cell(row=row, column=2).number_format = _NUM_FMT
    row += 1
    ws.cell(row=row, column=1, value="НДФЛ за год:")
    ws.cell(row=row, column=2, value=float(total_tax))
    ws.cell(row=row, column=2).number_format = _NUM_FMT
    row += 1
    ws.cell(row=row, column=1, value="Эффективная ставка:")
    ws.cell(row=row, column=2, value=f"{effective}%")
    row += 1
    ws.cell(row=row, column=1, value="После НДФЛ:")
    ws.cell(row=row, column=2, value=float(net))
    ws.cell(row=row, column=2).number_format = _NUM_FMT
    row += 1
    ws.cell(row=row, column=1, value="В месяц (после НДФЛ):")
    ws.cell(row=row, column=2, value=float((net / 12).quantize(TWO_PLACES, ROUND_HALF_UP)))
    ws.cell(row=row, column=2).number_format = _NUM_FMT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
